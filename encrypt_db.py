import os
import shutil
import tempfile
import json
import csv
import hashlib
import base64
import openpyxl

def rc4(key, data):
    S = list(range(256))
    j = 0
    out = []
    # KSA
    for i in range(256):
        j = (j + S[i] + key[i % len(key)]) % 256
        S[i], S[j] = S[j], S[i]
    # PRGA
    i = 0
    j = 0
    for byte in data:
        i = (i + 1) % 256
        j = (j + S[i]) % 256
        S[i], S[j] = S[j], S[i]
        t = (S[i] + S[j]) % 256
        k = S[t]
        out.append(byte ^ k)
    return bytes(out)

def encrypt(password, plaintext):
    key = hashlib.sha256(password.encode('utf-8')).digest()
    encrypted_bytes = rc4(key, plaintext.encode('utf-8'))
    return base64.b64encode(encrypted_bytes).decode('utf-8')

def parse_csv(csv_path):
    participantes = []
    with open(csv_path, mode='r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return participantes
        
        email_idx = None
        first_name_idx = None
        last_name_idx = None

        for idx, h in enumerate(headers):
            h_clean = str(h).strip().lower()
            if h_clean == 'email':
                email_idx = idx
            elif h_clean in ['first name', 'nombre']:
                first_name_idx = idx
            elif h_clean in ['last name', 'apellido', 'apellidos']:
                last_name_idx = idx

        if email_idx is None or first_name_idx is None:
            print(f"No se encontraron columnas necesarias en CSV {csv_path}")
            return participantes

        for row in reader:
            if len(row) > max(email_idx, first_name_idx):
                email = str(row[email_idx]).strip().lower()
                first_name = str(row[first_name_idx]).strip()
                last_name = str(row[last_name_idx]).strip() if last_name_idx is not None and len(row) > last_name_idx else ""
                
                if email and first_name:
                    participantes.append({
                        "first_name": first_name,
                        "last_name": last_name,
                        "email": email
                    })
    return participantes

def main():
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
    xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    participantes = []
    
    if csv_files:
        print(f"Procesando archivo CSV encontrado: '{csv_files[0]}'")
        participantes = parse_csv(csv_files[0])
    elif xlsx_files:
        excel_path = xlsx_files[0]
        print(f"Procesando archivo Excel encontrado: '{excel_path}'")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 10) if ws.cell(1, c).value is not None]
        
        email_idx = None
        first_name_idx = None
        last_name_idx = None

        for idx, h in enumerate(headers):
            h_clean = str(h).strip().lower()
            if h_clean == 'email':
                email_idx = idx + 1
            elif h_clean in ['first name', 'nombre']:
                first_name_idx = idx + 1
            elif h_clean in ['last name', 'apellido', 'apellidos']:
                last_name_idx = idx + 1

        if email_idx and first_name_idx:
            for row in range(2, ws.max_row + 1):
                email = ws.cell(row, email_idx).value
                first_name = ws.cell(row, first_name_idx).value
                last_name = ws.cell(row, last_name_idx).value if last_name_idx else ""

                if email and first_name:
                    participantes.append({
                        "first_name": str(first_name).strip(),
                        "last_name": str(last_name).strip() if last_name else "",
                        "email": str(email).strip().lower()
                    })

    print(f"Total de participantes extraídos: {len(participantes)}")
    if len(participantes) == 0:
        print("Error: No se encontraron registros válidos de participantes.")
        return

    password = "GoogleIO"
    out_filename = "participantes_io.rar"

    json_data = json.dumps(participantes, ensure_ascii=False)
    encrypted_str = encrypt(password, json_data)

    with open(out_filename, "w", encoding="utf-8") as f:
        f.write(encrypted_str)
    
    print(f"¡Éxito! Base de datos web cifrada generada en '{out_filename}' con la clave '{password}'")

if __name__ == "__main__":
    main()
